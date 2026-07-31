import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.exceptions import ValidationError
from app.schemas.device_contract import QUANTITY_UNITS, DeviceContractItem

# 与前端 ContractView 设备明细表头保持一致
COLUMN_SPECS: list[tuple[str, str, tuple[str, ...]]] = [
    ("item_kind", "类别", ("软硬类别", "类型")),
    ("device_name", "设备名称", ()),
    ("device_model_name", "设备型号", ()),
    ("manufacturer_name", "厂商", ()),
    ("quantity", "采购数量", ()),
    ("quantity_unit", "数量单位", ()),
    ("unit_price", "单价", ()),
    ("price_unit", "金额单位", ("单位(元/万元)", "单位")),
]

HEADER_LABELS = [spec[1] for spec in COLUMN_SPECS]
HEADER_KEYS = [spec[0] for spec in COLUMN_SPECS]
HEADER_ALIASES: dict[str, str] = {}
for key, label, aliases in COLUMN_SPECS:
    HEADER_ALIASES[label] = key
    for alias in aliases:
        HEADER_ALIASES[alias] = key

# 示例行：类别可选 硬件/软件；数量单位可选 台/个/件/套；金额单位可选 元/万元
SAMPLE_ROWS: list[list[object]] = [
    ["硬件", "服务器", "PowerEdge R750", "Dell", 10, "台", 8500, "元"],
    ["硬件", "核心交换机", "CE6857-48S6CQ-EI", "华为", 2, "套", 12, "万元"],
    ["软件", "操作系统许可", "Windows Server 2022", "Microsoft", 10, "套", 3200, "元"],
]

FIELD_HINTS = [
    "硬件 或 软件，默认硬件",
    "必填，最多100字",
    "必填，最多100字",
    "选填，最多100字",
    "非负整数",
    "台/个/件/套，默认台",
    "选填，支持小数",
    "元 或 万元，默认元",
]


def _normalize_item_kind(raw: object) -> str:
    text = str(raw or "").strip().lower()
    if text in ("software", "软", "软件", "许可", "license"):
        return "software"
    return "hardware"


def _normalize_price_unit(raw: object) -> str:
    text = str(raw or "").strip().lower()
    if text in ("wan", "万元", "万"):
        return "wan"
    return "yuan"


def _normalize_quantity_unit(raw: object) -> str:
    text = str(raw or "").strip()
    return text if text in QUANTITY_UNITS else "台"


def _to_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _build_header_index(header_row: list[str]) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        label = str(raw or "").strip()
        if not label:
            continue
        key = HEADER_ALIASES.get(label) or HEADER_ALIASES.get(label.lower())
        if key and key not in index_map:
            index_map[key] = idx
    return index_map


class ContractItemsExportService:
    def template_excel(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "设备明细"

        header_fill = PatternFill("solid", fgColor="F2F6FC")
        header_font = Font(bold=True)
        for col, label in enumerate(HEADER_LABELS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = max(len(label) * 2.2, 12)

        for row_idx, sample in enumerate(SAMPLE_ROWS, start=2):
            for col_idx, value in enumerate(sample, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        hint_ws = wb.create_sheet("填写说明")
        hint_ws.append(["列名", "说明"])
        for label, hint in zip(HEADER_LABELS, FIELD_HINTS, strict=True):
            hint_ws.append([label, hint])
        hint_ws.column_dimensions["A"].width = 14
        hint_ws.column_dimensions["B"].width = 28

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def parse_excel(self, content: bytes) -> tuple[list[DeviceContractItem], list[str]]:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError(f"无效的 Excel 文件: {exc}") from exc

        ws = wb["设备明细"] if "设备明细" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("Excel 文件为空")

        index_map = _build_header_index([str(cell or "").strip() for cell in rows[0]])
        if "device_name" not in index_map or "device_model_name" not in index_map:
            expected = "、".join(HEADER_LABELS[:2])
            raise ValidationError(f"缺少必要列，请使用最新模板（需包含：{expected}）")

        items: list[DeviceContractItem] = []
        errors: list[str] = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue

            def cell(key: str) -> object:
                idx = index_map.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            name = str(cell("device_name") or "").strip()
            model = str(cell("device_model_name") or "").strip()
            mfg = str(cell("manufacturer_name") or "").strip() or None
            qty_raw = cell("quantity")
            try:
                quantity = int(float(qty_raw)) if qty_raw not in (None, "") else 0
            except (TypeError, ValueError):
                errors.append(f"第 {row_num} 行：采购数量无效")
                continue
            quantity_unit = _normalize_quantity_unit(cell("quantity_unit"))
            unit_price = _to_decimal(cell("unit_price"))
            price_unit = _normalize_price_unit(cell("price_unit"))
            item_kind = _normalize_item_kind(cell("item_kind"))

            if not name and not model:
                continue
            if not name or not model:
                errors.append(f"第 {row_num} 行：设备名称与型号需同时填写")
                continue
            if quantity < 0:
                errors.append(f"第 {row_num} 行：采购数量不能为负数")
                continue

            items.append(
                DeviceContractItem(
                    device_name=name[:100],
                    device_model_name=model[:100],
                    manufacturer_name=mfg[:100] if mfg else None,
                    item_kind=item_kind,
                    quantity=min(quantity, 100000),
                    quantity_unit=quantity_unit,
                    unit_price=unit_price,
                    price_unit=price_unit,
                )
            )

        if not items and not errors:
            raise ValidationError("未解析到有效设备明细，请检查模板内容")
        return items, errors
