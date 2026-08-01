"""接口设计 Excel 导入导出。"""

from __future__ import annotations

import io
import uuid
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError, ValidationError
from app.schemas.export import ImportResult
from app.schemas.network import (
    CanvasLinkInput,
    CanvasNodeInput,
    CanvasSaveRequest,
    NetworkLinkType,
    NetworkTopologyDetailResponse,
    PortLayout,
)
from app.services.network import NetworkDesignService


class NetworkInterfaceExportService:
    EXPORT_HEADERS = [
        "场景",
        "连线类型",
        "本端类型",
        "本端名称",
        "本端位置",
        "本端U位",
        "本端接口ID",
        "本端接口",
        "对端类型",
        "对端名称",
        "对端位置",
        "对端U位",
        "对端接口ID",
        "对端接口",
        "接口类",
        "线缆类",
        "本端标签",
        "对端标签",
        "备注",
        "本端节点ID",
        "对端节点ID",
        "连线ID",
    ]

    ROLE_MAP = {
        "服务器接入": "server",
        "安全设备": "security",
        "上联": "uplink",
        "下联": "downlink",
        "互联": "interconnect",
        "server": "server",
        "security": "security",
        "uplink": "uplink",
        "downlink": "downlink",
        "interconnect": "interconnect",
    }
    TYPE_MAP = {
        "交换机-服务器": "switch_server",
        "交换机-交换机": "switch_switch",
        "交换机-安全": "switch_security",
        "switch_server": "switch_server",
        "switch_switch": "switch_switch",
        "switch_security": "switch_security",
    }
    CLASS_MAP = {
        "电口": "electric",
        "光口": "optical",
        "高速铜缆": "dac",
        "其他": "other",
        "electric": "electric",
        "optical": "optical",
        "dac": "dac",
        "other": "other",
    }
    CABLE_MAP = {
        "超六类铜缆": "copper_cat6",
        "多模光纤": "fiber_mm",
        "单模光纤": "fiber_sm",
        "DAC": "dac",
        "AOC": "aoc",
        "其他": "other",
        "copper_cat6": "copper_cat6",
        "fiber_mm": "fiber_mm",
        "fiber_sm": "fiber_sm",
        "dac": "dac",
        "aoc": "aoc",
        "other": "other",
    }
    ROLE_LABEL = {
        "server": "服务器接入",
        "security": "安全设备",
        "uplink": "上联",
        "downlink": "下联",
        "interconnect": "互联",
    }
    TYPE_LABEL = {
        "switch_server": "交换机-服务器",
        "switch_switch": "交换机-交换机",
        "switch_security": "交换机-安全",
    }
    CLASS_LABEL = {
        "electric": "电口",
        "optical": "光口",
        "dac": "高速铜缆",
        "other": "其他",
    }
    CABLE_LABEL = {
        "copper_cat6": "超六类铜缆",
        "fiber_mm": "多模光纤",
        "fiber_sm": "单模光纤",
        "dac": "DAC",
        "aoc": "AOC",
        "other": "其他",
    }

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.network = NetworkDesignService(session)

    def _node_map(self, detail: NetworkTopologyDetailResponse) -> dict[str, Any]:
        return {str(n.id): n for n in detail.nodes}

    def _port_label(self, node: Any, port_id: str) -> str:
        layout = getattr(node, "port_layout", None)
        if not layout or not getattr(layout, "ports", None):
            return port_id
        for p in layout.ports:
            if p.id == port_id:
                return p.label or port_id
        return port_id

    def _location(self, node: Any) -> tuple[str, str]:
        device = getattr(node, "device", None)
        if not device:
            return "", ""
        parts = []
        if device.room_name:
            parts.append(device.room_name)
        if device.rack_code:
            parts.append(device.rack_code)
        u = "" if device.u_position is None else str(device.u_position)
        return " / ".join(parts), u

    def _kind_label(self, node: Any) -> str:
        kind = getattr(node.kind, "value", node.kind)
        mapping = {"switch": "交换机", "server": "服务器", "security": "安全设备"}
        return mapping.get(str(kind), str(kind))

    async def export_excel(self, topology_id: uuid.UUID) -> bytes:
        detail = await self.network.get_detail(topology_id)
        nodes = self._node_map(detail)
        wb = Workbook()
        ws = wb.active
        ws.title = "接口设计"
        ws.append(self.EXPORT_HEADERS)
        for link in detail.links:
            source = nodes.get(str(link.source_node_id))
            target = nodes.get(str(link.target_node_id))
            s_loc, s_u = self._location(source) if source else ("", "")
            t_loc, t_u = self._location(target) if target else ("", "")
            role = link.link_role or ""
            ltype = getattr(link.link_type, "value", link.link_type)
            ws.append(
                [
                    self.ROLE_LABEL.get(role, role),
                    self.TYPE_LABEL.get(str(ltype), str(ltype)),
                    self._kind_label(source) if source else "",
                    source.name if source else "",
                    s_loc,
                    s_u,
                    link.source_port,
                    self._port_label(source, link.source_port) if source else link.source_port,
                    self._kind_label(target) if target else "",
                    target.name if target else "",
                    t_loc,
                    t_u,
                    link.target_port,
                    self._port_label(target, link.target_port) if target else link.target_port,
                    self.CLASS_LABEL.get(link.interface_class or "", link.interface_class or ""),
                    self.CABLE_LABEL.get(link.cable_type or "", link.cable_type or ""),
                    link.source_label or "",
                    link.target_label or "",
                    link.label or "",
                    str(link.source_node_id),
                    str(link.target_node_id),
                    str(link.id),
                ]
            )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def template_excel(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "接口设计"
        ws.append(self.EXPORT_HEADERS)
        ws.append(
            [
                "服务器接入",
                "交换机-服务器",
                "交换机",
                "示例交换机",
                "",
                "",
                "p1",
                "GE1",
                "服务器",
                "示例服务器",
                "",
                "",
                "slot1-p1",
                "NIC1",
                "电口",
                "超六类铜缆",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _cell(self, row: tuple, idx: int) -> str:
        if idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def _find_node(self, nodes: list[Any], node_id: str, name: str) -> Any | None:
        if node_id:
            for n in nodes:
                if str(n.id) == node_id:
                    return n
        if name:
            for n in nodes:
                if n.name == name:
                    return n
                device = getattr(n, "device", None)
                if device and (device.name == name or device.hostname == name):
                    return n
        return None

    def _find_port(self, node: Any, port_id: str, port_label: str) -> str | None:
        layout: PortLayout | None = getattr(node, "port_layout", None)
        ports = list(getattr(layout, "ports", None) or [])
        if port_id:
            for p in ports:
                if p.id == port_id:
                    return p.id
            # 允许直接使用端口 ID（无 port_layout 时）
            if not ports:
                return port_id
        if port_label:
            for p in ports:
                if (p.label or "") == port_label or p.id == port_label:
                    return p.id
        return port_id or None

    def _to_canvas_nodes(self, detail: NetworkTopologyDetailResponse) -> list[CanvasNodeInput]:
        result: list[CanvasNodeInput] = []
        for n in detail.nodes:
            result.append(
                CanvasNodeInput(
                    id=n.id,
                    kind=n.kind,
                    name=n.name,
                    device_id=n.device_id,
                    device_model_id=getattr(n, "device_model_id", None),
                    contract_device_name=getattr(n, "contract_device_name", None),
                    pos_x=n.pos_x,
                    pos_y=n.pos_y,
                    switch_port_count=n.switch_port_count,
                    slots=n.slots,
                    port_layout=n.port_layout,
                    on_canvas=bool(getattr(n, "on_canvas", True)),
                )
            )
        return result

    def _to_canvas_links(self, detail: NetworkTopologyDetailResponse) -> list[CanvasLinkInput]:
        result: list[CanvasLinkInput] = []
        for l in detail.links:
            result.append(
                CanvasLinkInput(
                    id=l.id,
                    link_type=l.link_type,
                    source_node_id=l.source_node_id,
                    source_port=l.source_port,
                    target_node_id=l.target_node_id,
                    target_port=l.target_port,
                    label=l.label,
                    source_label=l.source_label,
                    target_label=l.target_label,
                    cable_type=l.cable_type,
                    interface_class=l.interface_class,
                    link_role=l.link_role,
                )
            )
        return result

    async def import_excel(
        self,
        topology_id: uuid.UUID,
        content: bytes,
        user_id: uuid.UUID | None = None,
    ) -> ImportResult:
        detail = await self.network.get_detail(topology_id)
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise ValidationError("无法解析 Excel 文件", code=10004) from exc
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("Excel 为空", code=10004)
        header = [str(c or "").strip() for c in rows[0]]
        if "本端接口ID" not in header and "本端名称" not in header:
            raise ValidationError("Excel 表头不正确，请使用导入模板", code=10004)
        col = {name: i for i, name in enumerate(header)}

        def take(row: tuple, key: str) -> str:
            idx = col.get(key)
            if idx is None:
                return ""
            return self._cell(row, idx)

        canvas_nodes = self._to_canvas_nodes(detail)
        links = self._to_canvas_links(detail)
        by_id = {str(l.id): i for i, l in enumerate(links) if l.id}

        created = 0
        updated = 0
        failed = 0
        errors: list[str] = []

        for ridx, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                source = self._find_node(
                    detail.nodes, take(row, "本端节点ID"), take(row, "本端名称")
                )
                target = self._find_node(
                    detail.nodes, take(row, "对端节点ID"), take(row, "对端名称")
                )
                if not source or not target:
                    raise ValidationError("找不到本端或对端设备（请核对名称/节点ID）")
                source_port = self._find_port(source, take(row, "本端接口ID"), take(row, "本端接口"))
                target_port = self._find_port(target, take(row, "对端接口ID"), take(row, "对端接口"))
                if not source_port or not target_port:
                    raise ValidationError("找不到本端或对端接口")
                role_raw = take(row, "场景")
                type_raw = take(row, "连线类型")
                link_type = self.TYPE_MAP.get(type_raw) or (
                    "switch_server"
                    if role_raw in ("服务器接入", "server")
                    else "switch_security"
                    if role_raw in ("安全设备", "security")
                    else "switch_switch"
                )
                if link_type not in ("switch_server", "switch_switch", "switch_security"):
                    link_type = "switch_switch"
                link_role = self.ROLE_MAP.get(role_raw) or None
                iface = self.CLASS_MAP.get(take(row, "接口类")) or None
                cable = self.CABLE_MAP.get(take(row, "线缆类")) or None
                link_id_raw = take(row, "连线ID")
                link_uuid = None
                if link_id_raw:
                    try:
                        link_uuid = uuid.UUID(link_id_raw)
                    except ValueError:
                        link_uuid = None
                payload = CanvasLinkInput(
                    id=link_uuid,
                    link_type=NetworkLinkType(link_type),
                    source_node_id=source.id,
                    source_port=source_port,
                    target_node_id=target.id,
                    target_port=target_port,
                    label=take(row, "备注") or None,
                    source_label=take(row, "本端标签") or None,
                    target_label=take(row, "对端标签") or None,
                    cable_type=cable,
                    interface_class=iface,
                    link_role=link_role,
                )
                if link_uuid and str(link_uuid) in by_id:
                    links[by_id[str(link_uuid)]] = payload
                    updated += 1
                else:
                    # 同端点覆盖
                    replaced = False
                    for i, existing in enumerate(links):
                        if (
                            str(existing.source_node_id) == str(source.id)
                            and existing.source_port == source_port
                            and str(existing.target_node_id) == str(target.id)
                            and existing.target_port == target_port
                        ):
                            payload.id = existing.id
                            links[i] = payload
                            updated += 1
                            replaced = True
                            break
                    if not replaced:
                        links.append(payload)
                        created += 1
            except Exception as exc:  # noqa: BLE001
                failed += 1
                errors.append(f"第{ridx}行: {exc}")

        if created or updated:
            await self.network.save_canvas(
                topology_id,
                CanvasSaveRequest(nodes=canvas_nodes, links=links),
                user_id=user_id,
            )
        return ImportResult(created=created + updated, failed=failed, errors=errors[:50])
