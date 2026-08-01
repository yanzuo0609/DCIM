"""设备参数档案 Excel 导入/导出。

列设计面向资源统计：CPU / 内存 / 系统盘 / 数据盘（最多 3 组规格）。
导入时仅用非空单元格覆盖，便于补齐未填参数。
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.exceptions import ValidationError
from app.schemas.device import (
    PARAM_DATA_DISK_EXPORT_SLOTS,
    ParamCpuSpec,
    ParamDiskSpec,
    ParamMemorySpec,
    ParamProfilePayload,
    ParamRaidSpec,
    param_is_complete,
)

COLUMN_SPECS: list[tuple[str, str, tuple[str, ...]]] = [
    ("code", "编码", ("档案编码",)),
    ("name", "设备名称", ("名称", "采购设备名称")),
    ("source_device_model", "设备型号", ("型号",)),
    ("source_manufacturer", "厂商", ("制造商",)),
    ("cpu_cores", "CPU核心数", ("CPU核数", "核心数")),
    ("cpu_architecture", "CPU架构", ("架构",)),
    ("cpu_model", "CPU型号", ()),
    ("memory_size_gb", "内存GB", ("内存容量GB", "内存")),
    ("memory_ddr_type", "DDR类型", ("内存类型",)),
    ("memory_modules", "内存条数", ("条数",)),
    ("sys_size_gb", "系统盘容量GB", ("系统盘GB",)),
    ("sys_count", "系统盘数量", ()),
    ("sys_interface", "系统盘接口", ()),
    ("sys_media", "系统盘介质", ("系统盘类型",)),
]

for _i in range(1, PARAM_DATA_DISK_EXPORT_SLOTS + 1):
    COLUMN_SPECS.extend(
        [
            (f"data{_i}_size_gb", f"数据盘{_i}容量GB", ()),
            (f"data{_i}_count", f"数据盘{_i}数量", ()),
            (f"data{_i}_interface", f"数据盘{_i}接口", ()),
            (f"data{_i}_media", f"数据盘{_i}介质", ()),
        ]
    )

COLUMN_SPECS.extend(
    [
        ("fan_count", "风扇数量", ()),
        ("fan_model", "风扇型号", ()),
        ("psu_power_w", "电源功率W", ("电源W",)),
        ("raid_model", "RAID型号", ()),
        ("raid_params", "RAID参数", ()),
        ("description", "描述", ("备注",)),
        ("is_complete", "是否已完善", ("完善状态",)),
        ("missing_fields", "缺失字段", ()),
    ]
)

HEADER_LABELS = [spec[1] for spec in COLUMN_SPECS]
HEADER_KEYS = [spec[0] for spec in COLUMN_SPECS]
HEADER_ALIASES: dict[str, str] = {}
for key, label, aliases in COLUMN_SPECS:
    HEADER_ALIASES[label] = key
    for alias in aliases:
        HEADER_ALIASES[alias] = key

FIELD_HINTS = [
    "必填；导入匹配主键之一",
    "必填；与采购汇总设备名称对应",
    "选填；同名多型号时用于区分",
    "选填",
    "正整数，资源统计核心字段",
    "c86 或 arm",
    "选填",
    "数值(GB)，资源统计核心字段",
    "如 DDR4 / DDR5",
    "非负整数",
    "系统盘单盘容量(GB)，资源统计核心字段",
    "系统盘块数，默认1",
    "SATA/SAS/NVMe/PCIe/M.2/U.2",
    "ssd / hdd / nvme",
]
for _i in range(1, PARAM_DATA_DISK_EXPORT_SLOTS + 1):
    FIELD_HINTS.extend(
        [
            f"数据盘{_i}单盘容量(GB)",
            f"数据盘{_i}块数",
            f"数据盘{_i}接口",
            f"数据盘{_i}介质 ssd/hdd/nvme",
        ]
    )
FIELD_HINTS.extend(
    [
        "选填",
        "选填",
        "选填",
        "选填",
        "选填",
        "选填",
        "导出只读；导入忽略",
        "导出只读；导入忽略",
    ]
)

SAMPLE_ROW: list[object] = [
    "P-server-r750-demo",
    "服务器",
    "PowerEdge R750",
    "Dell",
    64,
    "c86",
    "Intel Xeon Gold 6338",
    512,
    "DDR4",
    16,
    480,
    2,
    "SATA",
    "ssd",
    1920,
    8,
    "SAS",
    "ssd",
    3840,
    4,
    "NVMe",
    "nvme",
    "",
    "",
    "",
    "",
    6,
    "标准风扇",
    1400,
    "PERC H755",
    "RAID10",
    "示例：完善后的服务器参数",
    "是",
    "",
]


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: object) -> int | None:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _to_float(value: object) -> float | None:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _normalize_arch(value: object) -> str | None:
    text = _cell_str(value).lower()
    if not text:
        return None
    if text in ("c86", "x86", "x86_64", "amd64"):
        return "c86"
    if text in ("arm", "aarch64", "arm64"):
        return "arm"
    return None


def _normalize_media(value: object) -> str | None:
    text = _cell_str(value).lower()
    if not text:
        return None
    if text in ("ssd", "固态", "solid"):
        return "ssd"
    if text in ("hdd", "机械", "机械盘"):
        return "hdd"
    if text in ("nvme",):
        return "nvme"
    return None


def _build_header_index(header_row: list[str]) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        label = str(raw or "").strip()
        if not label:
            continue
        key = HEADER_ALIASES.get(label) or HEADER_ALIASES.get(label.lower())
        if key and key not in index_map:
            index_map[key] = idx
    return index_map


def _disk_tuple(disk: ParamDiskSpec | None) -> tuple[Any, Any, Any, Any]:
    if not disk:
        return ("", "", "", "")
    return (
        disk.size_gb if disk.size_gb is not None else "",
        disk.count if disk.count is not None else "",
        disk.interface or "",
        disk.media_type or "",
    )


def profile_to_row(profile: dict[str, Any]) -> list[object]:
    payload_raw = profile.get("payload") or {}
    try:
        payload = (
            payload_raw
            if isinstance(payload_raw, ParamProfilePayload)
            else ParamProfilePayload.model_validate(payload_raw)
        )
    except Exception:  # noqa: BLE001
        payload = ParamProfilePayload()

    disks = list(payload.disks or [])
    system = next((d for d in disks if d.role == "system"), None)
    data_disks = [d for d in disks if d.role == "data"]
    if system is None and disks:
        # 兼容旧数据：首条视为系统盘
        system = disks[0]
        data_disks = disks[1:]

    row_map: dict[str, object] = {
        "code": profile.get("code") or "",
        "name": profile.get("name") or payload.source_device_name or "",
        "source_device_model": payload.source_device_model or "",
        "source_manufacturer": payload.source_manufacturer or "",
        "cpu_cores": payload.cpu.cores if payload.cpu else "",
        "cpu_architecture": payload.cpu.architecture if payload.cpu else "",
        "cpu_model": (payload.cpu.model if payload.cpu else None) or "",
        "memory_size_gb": payload.memory.size_gb if payload.memory else "",
        "memory_ddr_type": (payload.memory.ddr_type if payload.memory else None) or "",
        "memory_modules": payload.memory.modules if payload.memory else "",
        "fan_count": payload.fan_count if payload.fan_count is not None else "",
        "fan_model": payload.fan_model or "",
        "psu_power_w": payload.psu_power_w if payload.psu_power_w is not None else "",
        "raid_model": (payload.raid.model if payload.raid else None) or "",
        "raid_params": (payload.raid.params if payload.raid else None) or "",
        "description": profile.get("description") or "",
        "is_complete": "是" if param_is_complete(payload) else "否",
        "missing_fields": "、".join(profile.get("missing_fields") or []),
    }
    sys_vals = _disk_tuple(system)
    row_map["sys_size_gb"] = sys_vals[0]
    row_map["sys_count"] = sys_vals[1]
    row_map["sys_interface"] = sys_vals[2]
    row_map["sys_media"] = sys_vals[3]
    for i in range(PARAM_DATA_DISK_EXPORT_SLOTS):
        vals = _disk_tuple(data_disks[i] if i < len(data_disks) else None)
        row_map[f"data{i + 1}_size_gb"] = vals[0]
        row_map[f"data{i + 1}_count"] = vals[1]
        row_map[f"data{i + 1}_interface"] = vals[2]
        row_map[f"data{i + 1}_media"] = vals[3]

    return [row_map.get(key, "") for key in HEADER_KEYS]


def _merge_disk(
    existing: ParamDiskSpec | None,
    *,
    role: str,
    size_gb: float | None,
    count: int | None,
    interface: str | None,
    media: str | None,
) -> ParamDiskSpec | None:
    base = existing.model_copy() if existing else ParamDiskSpec(role=role)  # type: ignore[arg-type]
    base.role = role  # type: ignore[assignment]
    if size_gb is not None:
        base.size_gb = size_gb
    if count is not None:
        base.count = count
    if interface:
        base.interface = interface
    if media:
        base.media_type = media  # type: ignore[assignment]
    if (
        base.size_gb is None
        and base.count is None
        and not base.interface
        and not base.media_type
    ):
        return ParamDiskSpec(role=role)  # type: ignore[arg-type]
    return base


def row_to_payload_patch(
    row: dict[str, object], existing: ParamProfilePayload | None
) -> ParamProfilePayload:
    """用 Excel 非空字段合并到已有 payload。"""
    base = existing.model_copy(deep=True) if existing else ParamProfilePayload()

    if _cell_str(row.get("source_device_model")):
        base.source_device_model = _cell_str(row.get("source_device_model"))
    if _cell_str(row.get("source_manufacturer")):
        base.source_manufacturer = _cell_str(row.get("source_manufacturer"))
    if _cell_str(row.get("name")) and not base.source_device_name:
        base.source_device_name = _cell_str(row.get("name"))

    cpu = base.cpu.model_copy() if base.cpu else ParamCpuSpec()
    cores = _to_int(row.get("cpu_cores"))
    arch = _normalize_arch(row.get("cpu_architecture"))
    cpu_model = _cell_str(row.get("cpu_model")) or None
    if cores is not None:
        cpu.cores = cores
    if arch is not None:
        cpu.architecture = arch  # type: ignore[assignment]
    if cpu_model:
        cpu.model = cpu_model
    if cpu.cores is not None or cpu.architecture or cpu.model:
        base.cpu = cpu

    memory = base.memory.model_copy() if base.memory else ParamMemorySpec()
    mem_size = _to_float(row.get("memory_size_gb"))
    ddr = _cell_str(row.get("memory_ddr_type")) or None
    modules = _to_int(row.get("memory_modules"))
    if mem_size is not None:
        memory.size_gb = mem_size
    if ddr:
        memory.ddr_type = ddr
    if modules is not None:
        memory.modules = modules
    if memory.size_gb is not None or memory.ddr_type or memory.modules is not None:
        base.memory = memory

    disks = list(base.disks or [])
    system = next((d for d in disks if d.role == "system"), None)
    data_list = [d for d in disks if d.role == "data"]
    others = [d for d in disks if d.role not in ("system", "data")]
    if system is None and disks and not any(d.role for d in disks):
        system = disks[0]
        data_list = disks[1:]
        others = []

    system = _merge_disk(
        system,
        role="system",
        size_gb=_to_float(row.get("sys_size_gb")),
        count=_to_int(row.get("sys_count")),
        interface=_cell_str(row.get("sys_interface")) or None,
        media=_normalize_media(row.get("sys_media")),
    )

    merged_data: list[ParamDiskSpec] = []
    for i in range(PARAM_DATA_DISK_EXPORT_SLOTS):
        prev = data_list[i] if i < len(data_list) else None
        slot = _merge_disk(
            prev,
            role="data",
            size_gb=_to_float(row.get(f"data{i + 1}_size_gb")),
            count=_to_int(row.get(f"data{i + 1}_count")),
            interface=_cell_str(row.get(f"data{i + 1}_interface")) or None,
            media=_normalize_media(row.get(f"data{i + 1}_media")),
        )
        if slot and (
            slot.size_gb is not None
            or slot.count is not None
            or slot.interface
            or slot.media_type
            or i < len(data_list)
        ):
            # 保留已有空占位行，便于再次导出
            if (
                slot.size_gb is not None
                or slot.count is not None
                or slot.interface
                or slot.media_type
                or (prev is not None)
            ):
                merged_data.append(slot)

    # 超出导出槽位的已有数据盘保留
    if len(data_list) > PARAM_DATA_DISK_EXPORT_SLOTS:
        merged_data.extend(data_list[PARAM_DATA_DISK_EXPORT_SLOTS:])

    new_disks: list[ParamDiskSpec] = []
    if system is not None:
        new_disks.append(system)
    new_disks.extend(merged_data)
    new_disks.extend(others)
    base.disks = new_disks

    fan_count = _to_int(row.get("fan_count"))
    fan_model = _cell_str(row.get("fan_model")) or None
    psu = _to_float(row.get("psu_power_w"))
    if fan_count is not None:
        base.fan_count = fan_count
    if fan_model:
        base.fan_model = fan_model
    if psu is not None:
        base.psu_power_w = psu

    raid = base.raid.model_copy() if base.raid else ParamRaidSpec()
    raid_model = _cell_str(row.get("raid_model")) or None
    raid_params = _cell_str(row.get("raid_params")) or None
    if raid_model:
        raid.model = raid_model
    if raid_params:
        raid.params = raid_params
    if raid.model or raid.params:
        base.raid = raid

    return base


class ParamProfileExportService:
    def template_excel(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "设备参数"

        header_fill = PatternFill("solid", fgColor="F2F6FC")
        header_font = Font(bold=True)
        for col, label in enumerate(HEADER_LABELS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = max(len(label) * 1.8, 12)

        for col_idx, value in enumerate(SAMPLE_ROW, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        hint_ws = wb.create_sheet("填写说明")
        hint_ws.append(["列名", "说明"])
        for label, hint in zip(HEADER_LABELS, FIELD_HINTS, strict=True):
            hint_ws.append([label, hint])
        hint_ws.column_dimensions["A"].width = 18
        hint_ws.column_dimensions["B"].width = 42
        hint_ws.append([])
        hint_ws.append(["匹配规则", "优先按「编码」匹配；否则按「设备名称」与采购汇总一一对应"])
        hint_ws.append(["更新规则", "仅用非空单元格覆盖，空单元格保留系统已有值，便于补齐未填参数"])
        hint_ws.append(["资源核心字段", "CPU核心数、内存GB、系统盘容量、至少一组数据盘容量"])
        hint_ws.append(["同步说明", "设备参数列表按采购汇总「设备名称」预生成空待填项，不含具体参数值"])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_excel(self, profiles: list[dict[str, Any]], *, incomplete_only: bool = False) -> bytes:
        rows = profiles
        if incomplete_only:
            rows = [p for p in profiles if not p.get("is_complete")]

        wb = Workbook()
        ws = wb.active
        ws.title = "设备参数"
        header_fill = PatternFill("solid", fgColor="F2F6FC")
        header_font = Font(bold=True)
        for col, label in enumerate(HEADER_LABELS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = max(len(label) * 1.8, 12)

        for r_idx, profile in enumerate(rows, start=2):
            for c_idx, value in enumerate(profile_to_row(profile), start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        hint_ws = wb.create_sheet("填写说明")
        hint_ws.append(["列名", "说明"])
        for label, hint in zip(HEADER_LABELS, FIELD_HINTS, strict=True):
            hint_ws.append([label, hint])
        hint_ws.column_dimensions["A"].width = 18
        hint_ws.column_dimensions["B"].width = 42

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def parse_excel(self, content: bytes) -> tuple[list[dict[str, object]], list[str]]:
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValidationError("无法解析 Excel 文件") from exc
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c or "").strip() for c in next(rows_iter)]
        except StopIteration as exc:
            raise ValidationError("Excel 为空") from exc
        index_map = _build_header_index(header)
        if "code" not in index_map and "name" not in index_map:
            raise ValidationError("缺少「编码」或「设备名称」列")

        parsed: list[dict[str, object]] = []
        errors: list[str] = []
        for row_no, raw in enumerate(rows_iter, start=2):
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            row: dict[str, object] = {}
            for key, idx in index_map.items():
                row[key] = raw[idx] if idx < len(raw) else None
            code = _cell_str(row.get("code"))
            name = _cell_str(row.get("name"))
            if not code and not name:
                errors.append(f"第{row_no}行：编码与设备名称均为空，已跳过")
                continue
            row["_row_no"] = row_no
            parsed.append(row)
        return parsed, errors
