import io
import math
import uuid
from datetime import datetime

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ValidationError
from app.models.device import Device, DeviceStatus
from app.repositories.device import DeviceModelRepository, DeviceRepository
from app.repositories.rack import RackRepository
from app.schemas.export import ImportResult


class DeviceExportService:
    EXPORT_HEADERS = [
        "hostname",
        "serial_number",
        "model_code",
        "model_name",
        "height_u",
        "status",
        "rack_code",
        "u_position",
        "power",
        "weight",
        "description",
    ]
    IMPORT_HEADERS = ["hostname", "serial_number", "model_code", "description"]

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.device_repo = DeviceRepository(session)
        self.model_repo = DeviceModelRepository(session)
        self.rack_repo = RackRepository(session)

    async def _load_export_rows(self) -> list[list]:
        devices, _ = await self.device_repo.list_paginated(page=1, page_size=10000)
        rows: list[list] = []
        for item in devices:
            device = await self.device_repo.get_by_id_with_model(item.id)
            if not device:
                continue
            rack_code = ""
            if device.rack_id:
                rack = await self.rack_repo.get_by_id(device.rack_id)
                rack_code = rack.code if rack else ""
            rows.append([
                device.hostname,
                device.serial_number,
                device.model.code if device.model else "",
                device.model.name if device.model else "",
                device.height_u,
                device.status,
                rack_code,
                device.u_position or "",
                float(device.power) if device.power else "",
                float(device.weight) if device.weight else "",
                device.description or "",
            ])
        return rows

    async def export_excel(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Devices"
        ws.append(self.EXPORT_HEADERS)
        for row in await self._load_export_rows():
            ws.append(row)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def export_pdf(self) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("RackDCIM Pro - Device Inventory Report", styles["Title"]),
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
            Spacer(1, 12),
        ]
        data = [self.EXPORT_HEADERS] + await self._load_export_rows()
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#409eff")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
            ])
        )
        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()

    async def import_excel(self, content: bytes, user_id: uuid.UUID | None = None) -> ImportResult:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True)
        except Exception as exc:
            raise ValidationError(f"Invalid Excel file: {exc}") from exc

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ImportResult(failed=1, errors=["Empty file"])

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        missing = [h for h in self.IMPORT_HEADERS[:3] if h not in headers]
        if missing:
            return ImportResult(failed=1, errors=[f"Missing columns: {', '.join(missing)}"])

        idx = {h: headers.index(h) for h in self.IMPORT_HEADERS if h in headers}
        result = ImportResult()
        model_cache: dict[str, object] = {}

        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            try:
                hostname = str(row[idx["hostname"]]).strip()
                serial = str(row[idx["serial_number"]]).strip()
                model_code = str(row[idx["model_code"]]).strip()
                description = (
                    str(row[idx["description"]]).strip()
                    if "description" in idx and row[idx["description"]]
                    else None
                )
                if not hostname or not serial or not model_code:
                    raise ValueError("hostname, serial_number, model_code are required")

                if await self.device_repo.get_by_serial(serial):
                    raise ValueError(f"serial_number already exists: {serial}")
                if await self.device_repo.get_by_hostname(hostname):
                    raise ValueError(f"hostname already exists: {hostname}")

                if model_code not in model_cache:
                    model = await self.model_repo.get_by_code(model_code)
                    if not model:
                        raise ValueError(f"model_code not found: {model_code}")
                    model_cache[model_code] = model

                model = model_cache[model_code]
                device = Device(
                    name=hostname,
                    hostname=hostname,
                    serial_number=serial,
                    device_model_id=model.id,
                    height_u=model.height_u,
                    status=DeviceStatus.STOCK.value,
                    description=description,
                    created_by=user_id,
                    updated_by=user_id,
                )
                await self.device_repo.create(device)
                result.created += 1
            except Exception as exc:
                result.failed += 1
                result.errors.append(f"Row {row_num}: {exc}")

        return result

    def import_template_excel(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "DeviceImport"
        ws.append(self.IMPORT_HEADERS)
        ws.append(["srv-001", "SN-001", "R750-2U", "Optional description"])
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
